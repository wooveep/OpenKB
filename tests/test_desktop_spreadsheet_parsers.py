"""Focused XLS/XLSX import coverage for the Desktop Document IR boundary."""

from __future__ import annotations

import json
import sqlite3
import struct
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PillowImage

from openkb.desktop_import import DesktopTextImportService
from openkb.desktop_import_artifacts import DesktopImportError
from openkb.desktop_raw_assets import DesktopRawAssetService
from openkb.desktop_workspace import DesktopKnowledgeBaseRuntime


def test_xlsx_import_retains_sheets_cells_formula_cache_merge_and_source_image(tmp_path):
    kb_dir = tmp_path / "desktop-kb"
    source = tmp_path / "revenue.xlsx"
    source.write_bytes(_workbook_with_sheets_and_image())
    DesktopKnowledgeBaseRuntime().create(kb_dir)

    imported = DesktopTextImportService(kb_dir).import_text(source)

    assert imported.document.source_format == "xlsx"
    raw_path = kb_dir / "raw" / f"{imported.document.raw_asset_sha256}.xlsx"
    assert raw_path.read_bytes() == source.read_bytes()
    with sqlite3.connect(kb_dir / ".openkb" / "state.sqlite3") as connection:
        table_locator = json.loads(
            connection.execute(
                "SELECT locator_json FROM document_ir_blocks "
                "WHERE kind = 'table' AND locator_json LIKE '%Overview%'"
            ).fetchone()[0]
        )
        image_locator = json.loads(
            connection.execute("SELECT locator_json FROM source_images").fetchone()[0]
        )
        image_path = connection.execute("SELECT storage_path FROM source_images").fetchone()[0]
    assert table_locator["sheet"] == "Overview"
    assert table_locator["cell_range"] == "A1:C3"
    assert table_locator["header_row"] == 2
    assert table_locator["headers"] == ["Month", "Amount", "Double"]
    assert table_locator["merged_ranges"] == ["A1:C1"]
    assert table_locator["formula_cells"] == {"C3": {"formula": "=B3*2", "cached_value": None}}
    assert image_locator["sheet"] == "Overview"
    assert image_locator["cell_range"] == "D2"
    assert (kb_dir / image_path).read_bytes().startswith(b"\x89PNG")

    reader = DesktopRawAssetService(kb_dir).read_document(imported.document.document_id)
    assert "# Overview" in reader.content
    assert "# Details" in reader.content
    assert "=B3*2 [cached value unavailable]" in reader.content
    assert len(reader.source_images) == 1


def test_xls_import_retains_legacy_worksheet_values_with_visible_limits(tmp_path):
    kb_dir = tmp_path / "desktop-kb"
    source = tmp_path / "legacy.xls"
    source.write_bytes(_minimal_xls())
    DesktopKnowledgeBaseRuntime().create(kb_dir)

    imported = DesktopTextImportService(kb_dir).import_text(source)

    assert imported.document.source_format == "xls"
    raw_path = kb_dir / "raw" / f"{imported.document.raw_asset_sha256}.xls"
    assert raw_path.read_bytes() == source.read_bytes()
    with sqlite3.connect(kb_dir / ".openkb" / "state.sqlite3") as connection:
        table_locator = json.loads(
            connection.execute(
                "SELECT locator_json FROM document_ir_blocks WHERE kind = 'table'"
            ).fetchone()[0]
        )
    assert table_locator["sheet"] == "Sheet 1"
    assert table_locator["cell_range"] == "A1:A2"
    assert table_locator["parser_warnings"] == [
        "legacy_formula_text_unavailable",
        "legacy_source_images_unavailable",
    ]

    reader = DesktopRawAssetService(kb_dir).read_document(imported.document.document_id)
    assert "# Sheet 1" in reader.content
    assert "Hello" in reader.content
    assert "42.0" in reader.content
    assert "Legacy XLS compatibility" in reader.content


@pytest.mark.parametrize(
    ("suffix", "error_code"),
    ((".xlsx", "invalid_xlsx_document"), (".xls", "invalid_xls_document")),
)
def test_corrupt_spreadsheet_has_a_stable_import_failure(tmp_path, suffix, error_code):
    kb_dir = tmp_path / "desktop-kb"
    source = tmp_path / f"broken{suffix}"
    source.write_bytes(b"not a spreadsheet")
    DesktopKnowledgeBaseRuntime().create(kb_dir)

    with pytest.raises(DesktopImportError) as error:
        DesktopTextImportService(kb_dir).import_text(source)

    assert error.value.code == error_code


def test_xlsx_with_a_malformed_worksheet_has_a_stable_import_failure(tmp_path):
    kb_dir = tmp_path / "desktop-kb"
    source = tmp_path / "broken-worksheet.xlsx"
    source.write_bytes(_workbook_with_malformed_worksheet())
    DesktopKnowledgeBaseRuntime().create(kb_dir)

    with pytest.raises(DesktopImportError) as error:
        DesktopTextImportService(kb_dir).import_text(source)

    assert error.value.code == "invalid_xlsx_document"


def _workbook_with_sheets_and_image() -> bytes:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview.merge_cells("A1:C1")
    overview["A1"] = "Revenue"
    overview.append(["Month", "Amount", "Double"])
    overview.append(["June", 10, "=B3*2"])
    image_bytes = BytesIO()
    PillowImage.new("RGB", (1, 1), "blue").save(image_bytes, format="PNG")
    overview.add_image(OpenpyxlImage(BytesIO(image_bytes.getvalue())), "D2")

    details = workbook.create_sheet("Details")
    details.append(["Key", "Value"])
    details.append(["Owner", "OpenKB"])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _workbook_with_malformed_worksheet() -> bytes:
    workbook = Workbook()
    workbook.active["A1"] = "Broken"
    valid = BytesIO()
    workbook.save(valid)

    malformed = BytesIO()
    with (
        ZipFile(BytesIO(valid.getvalue())) as source,
        ZipFile(malformed, "w", compression=ZIP_DEFLATED) as destination,
    ):
        for item in source.infolist():
            content = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                content = b"<worksheet>"
            destination.writestr(item, content)
    return malformed.getvalue()


def _minimal_xls() -> bytes:
    def record(code: int, data: bytes = b"") -> bytes:
        return struct.pack("<HH", code, len(data)) + data

    bof = record(0x0009, struct.pack("<HH", 0, 0x0010))
    label = record(0x0004, struct.pack("<HH", 0, 0) + b"\0\0\0" + b"\x05Hello")
    number = record(0x0003, struct.pack("<HH", 1, 0) + b"\0\0\0" + struct.pack("<d", 42.0))
    return bof + label + number + record(0x000A)
