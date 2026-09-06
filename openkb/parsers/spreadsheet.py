"""Structure-preserving XLS and XLSX adapters for Desktop document import."""

from __future__ import annotations

from datetime import date, datetime, time
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import uuid4
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from xlrd import XLRDError, open_workbook
from xlrd.compdoc import CompDocError

from openkb.importing.artifacts import (
    DesktopImportError,
    DocumentIRBlock,
    ParsedDocument,
    SourceImage,
    source_image_from_content,
)


def parse_spreadsheet_document(
    source: Path, raw_bytes: bytes, source_format: str
) -> ParsedDocument:
    """Parse an XLS or XLSX Raw Asset into citable worksheet-level Document IR."""
    if source_format == "xlsx":
        return _parse_xlsx(source, raw_bytes)
    if source_format == "xls":
        return _parse_xls(source, raw_bytes)
    raise DesktopImportError(
        "unsupported_import_format", f"No spreadsheet parser is registered for {source.name}."
    )


def _parse_xlsx(source: Path, raw_bytes: bytes) -> ParsedDocument:
    formulas = None
    cached_values = None
    try:
        formulas = load_workbook(BytesIO(raw_bytes), data_only=False, read_only=False)
        cached_values = load_workbook(BytesIO(raw_bytes), data_only=True, read_only=False)
        blocks: list[DocumentIRBlock] = []
        images: list[SourceImage] = []
        for sheet_index, sheet in enumerate(formulas.worksheets, start=1):
            _append_xlsx_sheet(
                sheet,
                cached_values[sheet.title],
                sheet_index=sheet_index,
                blocks=blocks,
                images=images,
            )
        return ParsedDocument(tuple(blocks), tuple(images))
    except (
        BadZipFile,
        InvalidFileException,
        KeyError,
        OSError,
        ParseError,
        ValueError,
    ) as error:
        raise DesktopImportError(
            "invalid_xlsx_document", f"XLSX source cannot be read: {source.name}"
        ) from error
    finally:
        if formulas is not None:
            formulas.close()
        if cached_values is not None:
            cached_values.close()


def _parse_xls(source: Path, raw_bytes: bytes) -> ParsedDocument:
    try:
        workbook = open_workbook(file_contents=raw_bytes, formatting_info=True)
    except (CompDocError, OSError, ValueError, XLRDError) as error:
        raise DesktopImportError(
            "invalid_xls_document", f"XLS source cannot be read: {source.name}"
        ) from error

    try:
        blocks: list[DocumentIRBlock] = []
        for sheet_index in range(workbook.nsheets):
            _append_xls_sheet(
                workbook.sheet_by_index(sheet_index),
                sheet_index=sheet_index + 1,
                blocks=blocks,
            )
        return ParsedDocument(tuple(blocks), ())
    finally:
        workbook.release_resources()


def _append_xlsx_sheet(
    sheet: Any,
    cached_sheet: Any,
    *,
    sheet_index: int,
    blocks: list[DocumentIRBlock],
    images: list[SourceImage],
) -> None:
    bounds = _xlsx_bounds(sheet)
    sheet_name = str(sheet.title)
    heading_path = (sheet_name,)
    _append_sheet_heading(
        sheet_name,
        sheet_index=sheet_index,
        cell_range=_range_from_bounds(bounds) if bounds is not None else "A1",
        blocks=blocks,
    )
    if bounds is not None:
        values, formula_cells = _xlsx_values(sheet, cached_sheet, bounds)
        _append_sheet_table(
            sheet_name,
            sheet_index=sheet_index,
            bounds=bounds,
            heading_path=heading_path,
            values=values,
            merged_ranges=tuple(sorted(str(value) for value in sheet.merged_cells.ranges)),
            formula_cells=formula_cells,
            native_tables=_xlsx_native_tables(sheet),
            parser_warnings=(),
            blocks=blocks,
        )
    _append_xlsx_images(
        sheet,
        sheet_index=sheet_index,
        heading_path=heading_path,
        blocks=blocks,
        images=images,
    )


def _append_xls_sheet(sheet: Any, *, sheet_index: int, blocks: list[DocumentIRBlock]) -> None:
    bounds = _xls_bounds(sheet)
    sheet_name = str(sheet.name)
    heading_path = (sheet_name,)
    _append_sheet_heading(
        sheet_name,
        sheet_index=sheet_index,
        cell_range=_range_from_bounds(bounds) if bounds is not None else "A1",
        blocks=blocks,
    )
    warning = (
        "Legacy XLS compatibility: cached cell values are retained; original formula "
        "expressions and embedded images are unavailable."
    )
    blocks.append(
        DocumentIRBlock(
            block_id=_block_id(),
            ordinal=len(blocks),
            kind="paragraph",
            text=warning,
            heading_path=heading_path,
            line_start=1,
            line_end=1,
            locator={
                "sheet": sheet_name,
                "sheet_index": sheet_index,
                "parser_warnings": [
                    "legacy_formula_text_unavailable",
                    "legacy_source_images_unavailable",
                ],
            },
        )
    )
    if bounds is None:
        return
    _append_sheet_table(
        sheet_name,
        sheet_index=sheet_index,
        bounds=bounds,
        heading_path=heading_path,
        values=_xls_values(sheet, bounds),
        merged_ranges=_xls_merged_ranges(sheet),
        formula_cells={},
        native_tables=(),
        parser_warnings=(
            "legacy_formula_text_unavailable",
            "legacy_source_images_unavailable",
        ),
        blocks=blocks,
    )


def _append_sheet_heading(
    sheet_name: str, *, sheet_index: int, cell_range: str, blocks: list[DocumentIRBlock]
) -> None:
    blocks.append(
        DocumentIRBlock(
            block_id=_block_id(),
            ordinal=len(blocks),
            kind="heading",
            text=sheet_name,
            heading_path=(sheet_name,),
            line_start=1,
            line_end=1,
            locator={
                "sheet": sheet_name,
                "sheet_index": sheet_index,
                "cell_range": cell_range,
                "heading_level": 1,
            },
        )
    )


def _append_sheet_table(
    sheet_name: str,
    *,
    sheet_index: int,
    bounds: tuple[int, int, int, int],
    heading_path: tuple[str, ...],
    values: list[list[str]],
    merged_ranges: tuple[str, ...],
    formula_cells: dict[str, dict[str, object]],
    native_tables: tuple[dict[str, object], ...],
    parser_warnings: tuple[str, ...],
    blocks: list[DocumentIRBlock],
) -> None:
    header_index = _header_row_index(values)
    header_row = bounds[0] + header_index
    headers = values[header_index]
    text = _markdown_table(values, bounds, header_index)
    blocks.append(
        DocumentIRBlock(
            block_id=_block_id(),
            ordinal=len(blocks),
            kind="table",
            text=text,
            heading_path=heading_path,
            line_start=bounds[0],
            line_end=bounds[1],
            locator={
                "sheet": sheet_name,
                "sheet_index": sheet_index,
                "cell_range": _range_from_bounds(bounds),
                "header_row": header_row,
                "headers": headers,
                "merged_ranges": list(merged_ranges),
                "formula_cells": formula_cells,
                "native_tables": list(native_tables),
                "parser_warnings": list(parser_warnings),
            },
        )
    )


def _xlsx_bounds(sheet: Any) -> tuple[int, int, int, int] | None:
    coordinates = [
        (cell.row, cell.column)
        for row in sheet.iter_rows()
        for cell in row
        if _has_value(cell.value)
    ]
    for merged_range in sheet.merged_cells.ranges:
        coordinates.extend(
            (
                (merged_range.min_row, merged_range.min_col),
                (merged_range.max_row, merged_range.max_col),
            )
        )
    return _bounds_from_coordinates(coordinates)


def _xls_bounds(sheet: Any) -> tuple[int, int, int, int] | None:
    coordinates = [
        (row_index + 1, column_index + 1)
        for row_index in range(sheet.nrows)
        for column_index in range(sheet.ncols)
        if _has_value(sheet.cell_value(row_index, column_index))
    ]
    for row_start, row_end, column_start, column_end in getattr(sheet, "merged_cells", ()):
        coordinates.extend(((row_start + 1, column_start + 1), (row_end, column_end)))
    return _bounds_from_coordinates(coordinates)


def _bounds_from_coordinates(
    coordinates: list[tuple[int, int]],
) -> tuple[int, int, int, int] | None:
    if not coordinates:
        return None
    rows, columns = zip(*coordinates, strict=True)
    return min(rows), max(rows), min(columns), max(columns)


def _xlsx_values(
    sheet: Any, cached_sheet: Any, bounds: tuple[int, int, int, int]
) -> tuple[list[list[str]], dict[str, dict[str, object]]]:
    values: list[list[str]] = []
    formula_cells: dict[str, dict[str, object]] = {}
    for row_index in range(bounds[0], bounds[1] + 1):
        row: list[str] = []
        for column_index in range(bounds[2], bounds[3] + 1):
            cell = sheet.cell(row_index, column_index)
            if cell.data_type == "f":
                cached_value = cached_sheet.cell(row_index, column_index).value
                formula = _cell_text(cell.value)
                formula_cells[cell.coordinate] = {
                    "formula": formula,
                    "cached_value": _json_cell_value(cached_value),
                }
                if cached_value is None:
                    row.append(f"{formula} [cached value unavailable]")
                else:
                    row.append(f"{formula} [cached: {_cell_text(cached_value)}]")
            else:
                row.append(_cell_text(cell.value))
        values.append(row)
    return values, formula_cells


def _xls_values(sheet: Any, bounds: tuple[int, int, int, int]) -> list[list[str]]:
    return [
        [
            _cell_text(sheet.cell_value(row_index - 1, column_index - 1))
            for column_index in range(bounds[2], bounds[3] + 1)
        ]
        for row_index in range(bounds[0], bounds[1] + 1)
    ]


def _xlsx_native_tables(sheet: Any) -> tuple[dict[str, object], ...]:
    return tuple(
        {
            "name": str(table.name),
            "cell_range": str(table.ref),
            "header_row_count": int(table.headerRowCount),
        }
        for table in sheet.tables.values()
    )


def _xls_merged_ranges(sheet: Any) -> tuple[str, ...]:
    return tuple(
        _range_from_bounds((row_start + 1, row_end, column_start + 1, column_end))
        for row_start, row_end, column_start, column_end in getattr(sheet, "merged_cells", ())
    )


def _append_xlsx_images(
    sheet: Any,
    *,
    sheet_index: int,
    heading_path: tuple[str, ...],
    blocks: list[DocumentIRBlock],
    images: list[SourceImage],
) -> None:
    for image_index, image in enumerate(getattr(sheet, "_images", ()), start=1):
        try:
            content = image._data()
        except (AttributeError, OSError, ValueError):
            continue
        cell_range = _xlsx_image_cell_range(getattr(image, "anchor", None))
        filename = Path(str(getattr(image, "path", ""))).name or f"image-{image_index}.png"
        locator = {
            "sheet": str(sheet.title),
            "sheet_index": sheet_index,
            "cell_range": cell_range,
        }
        source_image = source_image_from_content(
            content=content,
            filename=filename,
            alt_text=f"{sheet.title} {cell_range}",
            ordinal=len(images),
            locator=locator,
        )
        images.append(source_image)
        blocks.append(
            DocumentIRBlock(
                block_id=_block_id(),
                ordinal=len(blocks),
                kind="figure",
                text=source_image.alt_text or source_image.filename,
                heading_path=heading_path,
                line_start=1,
                line_end=1,
                locator={**locator, "source_image_id": source_image.image_id},
            )
        )


def _xlsx_image_cell_range(anchor: object) -> str:
    if isinstance(anchor, str):
        return anchor
    start = _xlsx_anchor_coordinate(getattr(anchor, "_from", None))
    end = _xlsx_anchor_coordinate(getattr(anchor, "to", None))
    if start and end:
        return f"{start}:{end}"
    return start or "unknown"


def _xlsx_anchor_coordinate(marker: object) -> str | None:
    row = getattr(marker, "row", None)
    column = getattr(marker, "col", None)
    if type(row) is not int or type(column) is not int:
        return None
    return f"{get_column_letter(column + 1)}{row + 1}"


def _header_row_index(values: list[list[str]]) -> int:
    for index, row in enumerate(values):
        if sum(bool(value) for value in row) > 1:
            return index
    for index, row in enumerate(values):
        if any(row):
            return index
    return 0


def _markdown_table(
    values: list[list[str]], bounds: tuple[int, int, int, int], header_index: int
) -> str:
    preamble = [
        f"{_range_from_bounds((bounds[0] + index, bounds[0] + index, bounds[2], bounds[3]))}: "
        + " | ".join(value for value in row if value)
        for index, row in enumerate(values[:header_index])
        if any(row)
    ]
    headers = _markdown_row(values[header_index])
    table = [headers, "| " + " | ".join("---" for _ in values[header_index]) + " |"]
    table.extend(_markdown_row(row) for row in values[header_index + 1 :])
    return "\n\n".join(preamble + ["\n".join(table)])


def _markdown_row(row: list[str]) -> str:
    return "| " + " | ".join(_markdown_cell(value) for value in row) + " |"


def _markdown_cell(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", "<br>")


def _range_from_bounds(bounds: tuple[int, int, int, int]) -> str:
    start_row, end_row, start_column, end_column = bounds
    start = f"{get_column_letter(start_column)}{start_row}"
    end = f"{get_column_letter(end_column)}{end_row}"
    return start if start == end else f"{start}:{end}"


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def _json_cell_value(value: object) -> object:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def _has_value(value: object) -> bool:
    return value is not None and value != ""


def _block_id() -> str:
    return uuid4().hex
